import matplotlib.pyplot as plt
import numpy as np


def draw_bar(labels, quants):
    width = 0.4
    ind = np.linspace(0.5, 10.5, 7)
    # make a square figure
    fig = plt.figure(1)
    ax = fig.add_subplot(111)
    # Bar Plot
    colors = ['#F27970', '#BB9727', '#54B345', '#32B897', '#05B9E2', '#8983BF', '#C76DA2']
    ax.bar(ind - width / 2, quants, width, color=colors)
    # Set the ticks on x-axis
    ax.set_xticks(ind)
    ax.set_xticklabels(labels)
    # labels
    ax.set_xlabel('Methods')
    ax.set_ylabel('F1-score')
    # title
    # ax.set_title('Top 10 GDP Countries', bbox={'facecolor': '0.8', 'pad': 5})
    # plt.grid(True)
    plt.xticks(rotation=20)
    plt.show()
    plt.close()

def draw_bar_3D(labels, quants):
    x = np.array([1, 3, 5, 7, 9, 11, 13])
    y = np.array([1, 1, 1, 1, 1, 1, 1])
    z = quants
    ax = plt.subplot(projection='3d')  # 三维图形
    ax.set_xticklabels(labels)
    ax.set_xlabel('Methods')
    ax.set_ylabel('F1-score')

    for xx, yy, zz in zip(x, y, z):
        color = ['#F27970', '#BB9727', '#54B345', '#32B897', '#05B9E2', '#8983BF', '#C76DA2']
        ax.bar3d(
            xx,  # 每个柱的x坐标
            yy,  # 每个柱的y坐标
            0,  # 每个柱的起始坐标
            dx=1,  # x方向的宽度
            dy=0.5,  # y方向的厚度
            dz=zz,  # z方向的高度
            color=color)  # 每个柱的颜色

    plt.xticks(rotation=20)
    plt.show()


labels = ['ours', 'Resnet50', 'VGG16', 'MobileNetV2', 'InceptionV3', 'Xia et al.', 'Kumars']

quants = np.array([0.5792, 0.1634, 0.1831, 0.2184, 0.1739, 0.4838, 0.4582])

import openpyxl

wb = openpyxl.load_workbook('C:\\Users\\27966\\Desktop\\sewer_detection\\files\\draw_F1.xlsx')
ws = wb['Sheet1']

for i in range(7):
    ws.cell(row=i+2, column=1).value = labels[i]
    ws.cell(row=i+2, column=2).value = quants[i]

wb.save('C:\\Users\\27966\\Desktop\\sewer_detection\\files\\draw_F1.xlsx')
